from Employee import *
from tkinter import *

import xlrd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import utils


class FileHandler:

    @staticmethod
    def load_employees(app, file_loc):
        #print("Loading " + file_loc)
        # load the employees on file
        try:
            with open(file_loc, 'r') as previous_file:  # opens and closes file for you
                for line in previous_file:
                    employ_nam = StringVar()
                    #split from spaces
                    info = line.split()

                    employ_nam.set(info[0].strip())
                    newbie = Employee(name=employ_nam.get(),
                                      fname=info[1],
                                      salary=info[2],
                                      file=info[3])
                    try:
                        FileHandler.load_attendence(app, newbie)
                    except FileNotFoundError:
                        print("No attendance file")

                    app.new_person(employ_nam, newbie)
                    app.employee_counter += 1
        except FileNotFoundError:
            print("file not created yet!")

    @staticmethod # think of a way to do this and like be able to save just at the end
    def save(app):
        #print("Saving...")
        with open("./save.txt", 'w') as file:
            for i in range(len(app.displayed_employees) - 1):
                file.write(app.displayed_employees[i].get_name() + " " +
                           app.displayed_employees[i].get_filename() + " " +
                           str(app.displayed_employees[i].get_salary()) + " " +
                           app.displayed_employees[i].get_attendence_file() + "\n")

            file.write(app.displayed_employees[-1].get_name() + " " +
                       app.displayed_employees[-1].get_filename() + " " +
                       str(app.displayed_employees[-1].get_salary()) + " " +
                       app.displayed_employees[-1].get_attendence_file())

    @staticmethod
    def load_attendence(app, employee):
        loc = "./" + employee.get_filename() + "/attendance.xlsx"
        print("Opening: " + loc)
        rtn = []
        try:
            for i in range(3):
                workbook = xlrd.open_workbook(loc)
                sheet = workbook.sheet_by_index(0)
                rtn.append(sheet.cell_value(i, 1))

            employee.set_hours(rtn[0])
            employee.set_bonus(rtn[1])
            employee.set_deduction(rtn[2])
        except FileNotFoundError:
            print("No file found for: attendance")

    @staticmethod
    def edit_attendence(employee, hours, bonus, deduction):
        loc = "./" + employee.get_filename() + "/attendance.xlsx"
        employee.set_hours(hours.get("1.0", "end").strip())
        employee.set_bonus(bonus.get("1.0", "end").strip())
        employee.set_deduction(deduction.get("1.0", "end").strip())

        try:
            wb = load_workbook(loc)
            print("option 1")
            ws = wb["Sheet1"]
            input_hrs = ws.cell(1, 2)
            input_bonus = ws.cell(2, 2)
            input_deduction = ws.cell(3, 2)

            input_hrs.value = hours.get("1.0", "end").strip()
            input_bonus.value = bonus.get("1.0", "end").strip()
            input_deduction.value = deduction.get("1.0", "end").strip()

            wb.save(loc)
            wb.close()
        except utils.exceptions.InvalidFileException:
            print("option 2")
            wb = Workbook() #TODO: fix this option 2 nooooo :C
            wb.create_sheet("Sheet1")
            ws = wb["Sheet1"]
            input_hrs = ws.cell(1, 2)
            input_bonus = ws.cell(2, 2)
            input_deduction = ws.cell(3, 2)

            input_hrs.value = hours.get("1.0", "end").strip()
            input_bonus.value = bonus.get("1.0", "end").strip()
            input_deduction.value = deduction.get("1.0", "end").strip()

            wb.save(loc)
            wb.close()

        return loc

